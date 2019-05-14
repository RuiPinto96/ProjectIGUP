import openpyxl
import sys
import time
import datetime
import calendar
from ConfigLimits import temp_max, temp_min, pressure_max, pressure_min, vapor_max, vapor_min, humidity_max, humidity_min, ozone_max, ozone_min, cloud_max, cloud_min, pluv_max, pluv_min, udo_max, udo_min, abso_wind_speed_max, abso_wind_speed_min, clock_wind_speed_max, clock_wind_speed_min, HgVapor_max, HgVapor_min, insolation_max, insolation_min, insolationPercentage_max, insolationPercentage_min


from os import listdir
from os.path import isfile, join
onlyfiles = [f for f in listdir('HumidadeRel') if isfile(join('HumidadeRel', f))]

from openpyxl import load_workbook
from openpyxl.workbook import Workbook


print(onlyfiles)


for file in onlyfiles:
    print('')
    print('-----------------------------------------------------')
    print('|       ' + '                                            |')
    print('       ' + file + '       ')
    print('|       ' + '                                            |')
    
    print('-----------------------------------------------------')
    
    file='HumidadeRel/' + file
    wb=load_workbook(filename=str(file), data_only=True)



    number_sheets= len(wb.sheetnames)

    #print(number_sheets)

    for index3,ws in enumerate(wb.worksheets):
        if(index3 < 12):
            if(ws==wb.worksheets[0]):
                a = file[-9:]
                ano = a[:4]
                ano_ver=int(ano)
            
            
        #FEVEREIRO
        if ws==wb.worksheets[1]:
            if(calendar.isleap(ano_ver)):
                minr=12
                maxc=31
                maxr=40
                minc=2
            else:
                minr=12
                maxc=31
                maxr=39
                minc=2
                
                
        #Mes com 31 dias
        if ws==wb.worksheets[0] or ws==wb.worksheets[2] or ws==wb.worksheets[4] or ws==wb.worksheets[6] or ws==wb.worksheets[7] or ws==wb.worksheets[9] or  ws==wb.worksheets[11]:
            minr=12
            maxc=31
            maxr=42
            minc=2
           
           
        #Mes com 30 dias    
        if ws==wb.worksheets[3] or ws==wb.worksheets[5] or ws==wb.worksheets[8] or ws==wb.worksheets[10]:
            minr=12
            maxc=31
            maxr=41
            minc=2
            
        
        
            
            
        if index3 <= 11:
            for index2,row in enumerate(ws.iter_rows(min_row=minr, max_col=maxc, max_row=maxr, min_col=minc)):
                for index, cell in enumerate(row):
                    dia=index2 + 1
                    mes=index3 + 1
                    hora = index + 1
                    #valores
                    if(index<24 and index>=0):
                        try:
                            value=float(cell.value)
                            
                            if(value > humidity_max or value < humidity_min):
                                print('Valor Suspeito Humidade(%): ' + str(value) + ' - ' + str(ws) + ' Dia: ' + str(dia) + ' Hora:' + str(hora))
                            
                        except ValueError:
                            print('Formato de Valor incorrecto: ' + str(ws) + ' Dia: ' + str(dia) + ' Hora :' + str(hora) + ' Coluna ' + str(index))
                        except TypeError:
                            print('Celula vazia: ' + str(ws) + ' Dia: ' + str(dia) + ' Hora:'  + str(hora))
                            pass
                        except AttributeError:
                            pass
                        if hora==24:
                            hora=0
                        #print(cell.value,int(ano),mes,dia,hora)
                        #formated_timestamp = (ano) + ',' + str(mes) + ',' + str(dia) + ',' + str(hora)
                        #dt = datetime.datetime.strptime(formated_timestamp, '%Y,%m,%d,%H')
                        #ut = time.mktime(dt.timetuple())
                        #print ut
                    #medias
                    if(index>=24):   
                        try:
                            value=float(cell.value)
                            
                            if(value > humidity_max or value < humidity_min):
                                print('Valor Suspeito Humidade(%): ' + str(value) + ' - ' + str(ws) + ' Dia: ' + str(dia) + ' Hora:' + str(hora))
                            
                        except ValueError:
                            print('Formato de Valor incorrecto: ' + str(ws) + ' Dia: ' + str(dia)  + ' Coluna ' + str(index))
                        except TypeError:
                            print('Celula vazia: ' + str(ws) + ' Dia: ' + str(dia) + ' Hora:'  + str(hora) + ' Coluna ' + str(index))
                            pass
                        except AttributeError:
                            pass
                        
                        #print(cell.value,int(ano),mes,dia)
                        #formated_timestamp = (ano) + ',' + str(mes) + ',' + str(dia)
                        #dt = datetime.datetime.strptime(formated_timestamp, '%Y,%m,%d')
                        #ut = time.mktime(dt.timetuple())
                        #print ut
    
    print(file)                    
    print('----------------------------')                        
    
