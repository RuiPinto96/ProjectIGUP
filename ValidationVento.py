
import sys
import time
import datetime
import calendar
from os import listdir
from os.path import isfile, join
onlyfiles = [f for f in listdir('VelocidadeVento') if isfile(join('VelocidadeVento', f))]

from openpyxl import load_workbook
from openpyxl.workbook import Workbook

print(onlyfiles)


for file in onlyfiles:
    file='VelocidadeVento/' + file
    wb=load_workbook(filename=str(file), data_only=True)



    number_sheets= len(wb.sheetnames)

    #print(number_sheets)

 

    for index3,ws in enumerate(wb.worksheets):
        if(index3 < 12):
            if(ws==wb.worksheets[0]):
                a = file[-9:]
                ano = a[:4]
                #print(ano)
                #cell83=False
                #if ano!=int:
                 #   cell83=True
                    #ano = str(ws.cell(row=5, column=2).value)
                ano_ver=int(ano)
            
            
        #FEVEREIRO
        if ws==wb.worksheets[1]:
            if(calendar.isleap(ano_ver)):
                minr=12
                maxc=31
                maxr=40
                minc=2
            else:
                minr=12
                maxc=31
                maxr=39
                minc=2
                
                
        #Mes com 31 dias
        if ws==wb.worksheets[0] or ws==wb.worksheets[2] or ws==wb.worksheets[4] or ws==wb.worksheets[6] or ws==wb.worksheets[7] or ws==wb.worksheets[9] or  ws==wb.worksheets[11]:
            minr=12
            maxc=31
            maxr=42
            minc=2
           
           
        #Mes com 30 dias    
        if ws==wb.worksheets[3] or ws==wb.worksheets[5] or ws==wb.worksheets[8] or ws==wb.worksheets[10]:
            minr=12
            maxc=31
            maxr=41
            minc=2
            
        
        
            
            
        if index3 <= 11:
            for index2,row in enumerate(ws.iter_rows(min_row=minr, max_col=maxc, max_row=maxr, min_col=minc)):
                for index, cell in enumerate(row):
                    dia=index2 + 1
                    mes=index3 + 1
                    hora = index + 1
                    #valores
                    if(index<24 and index>=0):
                        try:
                            value=float(cell.value)
                            if("a" in cell.value):
                                print('Avaria')
                            if("vest" in cell.value):
                                print('vestigios')
                        except ValueError:
                            print('Valores Errados:')
                            print('Mes: ' + str(ws) + ' Dia: ' + str(dia) + ' Hora :' + str(hora))
                        except TypeError:
                            pass
                            #print('Celula vazia')
                        except AttributeError:
                            pass
                        if hora==24:
                            hora=0
                        #print(cell.value,int(ano),mes,dia,hora)
                        formated_timestamp = str(ano) + ',' + str(mes) + ',' + str(dia) + ',' + str(hora)
                        dt = datetime.datetime.strptime(formated_timestamp, '%Y,%m,%d,%H')
                        ut = time.mktime(dt.timetuple())
                        #print ut
                    #medias
                    elif(index>=24): 
                        try:
                            
                            if("a" in cell):
                                print('Avaria')
                            if(index!=27):    
                                value=float(cell.value)
                            if(index==27):
                                value=str(cell.value)
                        except ValueError:
                            print('Valores Errados:')
                            print('Mes: ' + str(ws) + ' Dia: ' + str(dia))                       
                        except TypeError:
                            pass
                            #print('Celula vazia')
                        except AttributeError:
                            pass
                        print(cell.value,int(ano),mes,dia)
                        formated_timestamp = str(ano) + ',' + str(mes) + ',' + str(dia)
                        dt = datetime.datetime.strptime(formated_timestamp, '%Y,%m,%d')
                        ut = time.mktime(dt.timetuple())
                        #print ut
                        
    print(file)                    
    print('----------------------------')                        
    
