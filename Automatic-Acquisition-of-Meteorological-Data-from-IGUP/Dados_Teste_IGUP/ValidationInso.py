
import sys
import time
import datetime
import calendar
from os import listdir
from os.path import isfile, join
onlyfiles = [f for f in listdir('Insolation') if isfile(join('Insolation', f))]

from openpyxl import load_workbook
from openpyxl.workbook import Workbook

print(onlyfiles)


for file in onlyfiles:
    file='Insolation/' + file
    wb=load_workbook(filename=str(file), data_only=True)



    number_sheets= len(wb.sheetnames)

    #print(number_sheets)

 

    for index3,ws in enumerate(wb.worksheets):
        if(index3 < 12):
            if(ws==wb.worksheets[0]):
                a = file[-9:]
                ano = a[:4]
                #print(ano)
                #cell83=False
                #if ano!=int:
                 #   cell83=True
                    #ano = str(ws.cell(row=5, column=2).value)
                ano_ver=int(ano)
            
            
        #FEVEREIRO
        if ws==wb.worksheets[1]:
            if(calendar.isleap(ano_ver)):
                minr=10
                maxc=32
                maxr=38
                minc=4
            else:
                minr=10
                maxc=32
                maxr=37
                minc=4
                
                
        #Mes com 31 dias
        if ws==wb.worksheets[0] or ws==wb.worksheets[2] or ws==wb.worksheets[4] or ws==wb.worksheets[6] or ws==wb.worksheets[7] or ws==wb.worksheets[9] or  ws==wb.worksheets[11]:
            minr=10
            maxc=32
            maxr=40
            minc=4
           
           
        #Mes com 30 dias    
        if ws==wb.worksheets[3] or ws==wb.worksheets[5] or ws==wb.worksheets[8] or ws==wb.worksheets[10]:
            minr=10
            maxc=32
            maxr=39
            minc=4
            
        
        
            
            
        if index3 <= 11:
            for index2,row in enumerate(ws.iter_rows(min_row=minr, max_col=maxc, max_row=maxr, min_col=minc)):
                for index, cell in enumerate(row):
                    dia=index2 + 1
                    mes=index3 + 1
                    hora = index + 4
                    #valores
                    if(index<17 and index>=0):
                        try:
                            value=float(cell.value)
                            if("a" in cell.value):
                                print('Avaria')
                            if("vest" in cell.value):
                                print('vestigios')
                        except ValueError:
                            print('Valores Errados:')
                            print('Mes: ' + str(ws) + ' Dia: ' + str(dia) + ' Hora :' + str(hora))
                        except TypeError:
                            pass
                            #print('Celula vazia')
                        except AttributeError:
                            pass
                        if hora==24:
                            hora='00'
                        #print(cell.value,int(ano),mes,dia,hora)
                        formated_timestamp = str(ano) + ',' + str(mes) + ',' + str(dia) + ',' + str(hora)
                        dt = datetime.datetime.strptime(formated_timestamp, '%Y,%m,%d,%H')
                        ut = time.mktime(dt.timetuple())
                        #print ut
                    #medias
                    elif(index>=22): 
                        try:
                            if(index==22):
                                cell = str(ws.cell(row=index2 + 12 , column=26).value) + ':' + str(ws.cell(row=index2 + 12, column=27).value)
                            if(index==24):
                                cell = str(ws.cell(row=index2 + 12, column=28).value) + ':' + str(ws.cell(row=index2 + 12, column=29).value)
                            if("a" in cell):
                                print('Avaria')
                            if(index!=22 and index!=24):    
                                value=float(cell.value)
                        except ValueError:
                            print('Valores Errados:')
                            print('Mes: ' + str(ws) + ' Dia: ' + str(dia))                       
                        except TypeError:
                            pass
                            #print('Celula vazia')
                        except AttributeError:
                            pass
                        #if(index!=23 and index!=25):    
                            #print(cell.value,int(ano),mes,dia)
                        formated_timestamp = str(ano) + ',' + str(mes) + ',' + str(dia)
                        dt = datetime.datetime.strptime(formated_timestamp, '%Y,%m,%d')
                        ut = time.mktime(dt.timetuple())
                        #print ut
                        
    #if(Wrong_year==True):                    
        #print('Ano esta mal! ')                    
    print(file)                    
    print('----------------------------')                        
    
