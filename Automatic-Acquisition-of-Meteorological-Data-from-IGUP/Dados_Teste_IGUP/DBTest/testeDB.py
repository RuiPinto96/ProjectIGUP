from influxdb import InfluxDBClient
import sys
import time
import datetime
import calendar
from os import listdir
from os.path import isfile, join
#onlyfiles = [f for f in listdir('TempAr') if isfile(join('TempAr', f))]

from openpyxl import load_workbook

#print(onlyfiles)


#for file in onlyfiles:
    #file='TempAr/' + file
def main():
    wb=load_workbook(filename='tempar2001.xlsx', data_only=True)



    number_sheets= len(wb.sheetnames)

    #print(number_sheets)



    for index3,ws in enumerate(wb.worksheets):
        if(index3 < 12):
            if(ws==wb.worksheets[0]):
                ano = str(ws.cell(row=6, column=2).value)
                cell83=False
                if ano=='None':
                    cell83=True
                    ano = str(ws.cell(row=8, column=3).value)
                ano_ver=int(ano)
            
            
        #FEVEREIRO
        if ws==wb.worksheets[1]:
            if(calendar.isleap(ano_ver)):
                minr=12
                maxc=31
                maxr=40
                minc=2
            else:
                minr=12
                maxc=31
                maxr=39
                minc=2
                
                
        #Mes com 31 dias
        if ws==wb.worksheets[0] or ws==wb.worksheets[2] or ws==wb.worksheets[4] or ws==wb.worksheets[6] or ws==wb.worksheets[7] or ws==wb.worksheets[9] or  ws==wb.worksheets[11]:
            minr=12
            maxc=31
            maxr=42
            minc=2
           
           
        #Mes com 30 dias    
        if ws==wb.worksheets[3] or ws==wb.worksheets[5] or ws==wb.worksheets[8] or ws==wb.worksheets[10]:
            minr=12
            maxc=31
            maxr=41
            minc=2
            
        if index3 <= 11:
            for index2,row in enumerate(ws.iter_rows(min_row=minr, max_col=maxc, max_row=maxr, min_col=minc, values_only=True)):
                for index, cell in enumerate(row):
                    if (ano!=str(ws.cell(row=6, column=2).value) and cell83==False) or (ano!=str(ws.cell(row=8, column=3).value) and cell83==True):
                        sys.exit('Ano esta mal! ' + file)
                    dia=index2 + 1
                    mes=index3 + 1
                    hora = index + 1
                    #valores
                    if(index<24 and index>=0):
                        #print(cell,int(ano),mes,dia,hora)
                        if hora==24:
                            hora='00'
                        formated_timestamp = (ano) + ',' + str(mes) + ',' + str(dia) + ',' + str(hora)
                        dt = datetime.datetime.strptime(formated_timestamp, '%Y,%m,%d,%H')
                        ut = time.mktime(dt.timetuple())
                        #print ut
                        bd(ut,cell)
                    #medias
                    else:   
                        #print(cell,int(ano),mes,dia)
                        formated_timestamp = (ano) + ',' + str(mes) + ',' + str(dia)
                        dt = datetime.datetime.strptime(formated_timestamp, '%Y,%m,%d')
                        ut = time.mktime(dt.timetuple())
                        #print ut
                        bd(ut,cell)
                        
                        
                        




def bd(ut, cell, host='localhost', port=8086):
    """Instantiate a connection to the InfluxDB."""
    user = 'root'
    password = 'root'
    dbname = 'IGUP'
    dbuser = 'smly'
    dbuser_password = 'my_secret_password'
    #query = 'select * from load'
    json_body = [
        {
            "measurement": "Temperature",
            "tags": {
                "Instrument": "Thermometer",
                "Units": "Celsius"
            },
            "time": int(ut),
            "fields": {
                "Value": float(cell)
            }
        }
    ]

    client = InfluxDBClient(host, port, user, password, dbname)

    
    #print("Write points: {0}".format(json_body))
    client.write_points(json_body)

    #print("Querying data: " + query)
    #result = client.query(query)

    #print("Result: {0}".format(result))
    
    

if __name__ == '__main__':
    main()
    
    
